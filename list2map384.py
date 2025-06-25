#!/usr/bin/python3
# list2map384.py: generate a 16x24 map (XLSX) from a list of 384 indices and associated identities (CSV)
# SLD 2025

# License

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# Imports
import csv, sys
from openpyxl import load_workbook
import openpyxl

# Template definitions
template_filename = "./template.xlsx"
output_row_origin = 1
output_col_origin = 1

# Output definitions
output_prefix = "MAP__"

# Input index labels
cmpd_id_label = "CompoundBatch"
plt_id_label = "Plate Barcode"
well_id_label = "Position"

# Announce the script
print()
print("list2map384.py: generate a 16x24 map (XLSX) from a list of 384 indices and associated identities (CSV)")
print("SLD 2025")
print()
print("Usage    :   python3 list2map384.py INPUT_CSV_FILENAME")
print()

# Parse command line arguments
# Expect exactly one: input CSV filename
# Announce and quit if we don't get one
try: csv_input_filename = sys.argv[1]
except: quit("ERROR: no input CSV filename specified\nQuitting immediately\n\n")

# If we're here, we have a valid input CSV filename
# From this, derive an appropriate output XLSX filename
output_filename = output_prefix+csv_input_filename[:-4]+".xlsx"

# Open CSV input file
# If this fails (e.g., if file doesn't exist), announce and quit
try: csvfile = open(csv_input_filename, "r+", encoding="utf-8-sig")
except: quit("ERROR: supplied input CSV filename "+csv_input_filename+" does not exist\nQuitting immediately\n\n")

# Ready to read data from input CSV file
# Prepare to determine list indices from the label row
cmpd_id_index = -1
plt_id_index = -1
well_id_index = -1

# Prepare a list of lists to receive input data as a 16x24 matrix
matrix = [[[""] for col_iterator in range(0,24)] for row_iterator in range(0,16)]

# Open CSV input file
reader = csv.reader(csvfile)
row_counter = 0
for this_row in reader:

    # If we're currently on the 0th row, this is the label row
    if row_counter == 0:

        # Define important list indices for constructing platemap matrix
        index_counter = 0
        for this_label in this_row:
            if this_label == cmpd_id_label: cmpd_id_index = index_counter
            if this_label == plt_id_label: plt_id_index = index_counter
            if this_label == well_id_label: well_id_index = index_counter
            index_counter += 1
        
        # Check list indices: if any are missing, announce and quit
        if cmpd_id_index == -1: quit("ERROR: compound index (labeled '"+cmpd_id_label+"') not defined\nQuitting immediately\n\n")
        if plt_id_index == -1: quit("ERROR: plate index (labeled '"+plt_id_label+"') not defined\nQuitting immediately\n\n")
        if well_id_index == -1: quit("ERROR: well index (labeled '"+well_id_label+"') not defined\nQuitting immediately\n\n")

    # If still running after the 0th row, this row contains data
    else:

        # Get data for this row (corresponds to 1w@384wp)
        this_cmpd_id = this_row[cmpd_id_index]
        this_plt_id = this_row[plt_id_index]
        this_well_id = this_row[well_id_index]

        # Process well ID into corresponding col/row pair in the matrix
        this_well_row = ord(this_well_id[0]) - ord("A")
        this_well_col = int(this_well_id[1:]) - 1

        # Set matrix value for this well ID to the relevant compound ID
        matrix[this_well_row][this_well_col] = this_cmpd_id

        # Report 
        #print("compound id  :", this_cmpd_id)
        #print("plate id     :", this_plt_id)
        #print("well id      :", this_well_row,":",this_well_col)
        #print()
        
    # Iterate row counter
    row_counter += 1

# Close CSV input file
csvfile.close()


# Open template XLSX file and select the appropriate sheet
wb = load_workbook(template_filename)
ws = wb["Sheet1"]

# Fill template platemap with matrix values
for matrix_row in range(0,16):
    for matrix_col in range(0,24):
        ws.cell(row=matrix_row+2, column=matrix_col+2, value = matrix[matrix_row][matrix_col])

# Set plate label with last detected plate ID
# We assume the input file lists the well contents (even if null) of a single 384w plate
ws.cell(row=19, column=2).value = this_plt_id

# Write output file
wb.save(output_filename)

# If we've gotten this far, announce as much
quit("Successfully parsed "+csv_input_filename+"\nSuccessfully written "+output_filename+"\nQuitting immediately\n\n")